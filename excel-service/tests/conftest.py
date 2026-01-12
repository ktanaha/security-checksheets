"""pytestの共通設定とフィクスチャ"""

import os
import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


@pytest.fixture
def temp_excel_file(tmp_path):
    """テスト用の一時Excelファイルを作成する"""
    file_path = tmp_path / "test_sample.xlsx"

    wb = Workbook()

    # Sheet1: シンプルなデータ
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1['A1'] = "質問"
    ws1['B1'] = "回答"
    ws1['A2'] = "個人情報の取扱いについて適切な管理体制が構築されていますか？"
    ws1['B2'] = "はい、個人情報保護方針に基づき適切に管理しています。"
    ws1['A3'] = "アクセス制御は適切に実施されていますか？"
    ws1['B3'] = "はい、役割ベースのアクセス制御（RBAC）を実施しています。"

    # Sheet2: セル結合を含むデータ
    ws2 = wb.create_sheet("Sheet2")
    ws2['A1'] = "部門"
    ws2['B1'] = "質問"
    ws2['C1'] = "回答"
    ws2.merge_cells('A2:A3')  # A2とA3を結合
    ws2['A2'] = "情報システム"
    ws2['B2'] = "バックアップは定期的に実施されていますか？"
    ws2['C2'] = "はい、毎日実施しています。"
    ws2['B3'] = "災害対策は講じられていますか？"
    ws2['C3'] = "はい、BCP計画を策定しています。"

    # Sheet3: 空のシート
    wb.create_sheet("Sheet3")

    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def large_excel_file(tmp_path):
    """大量データを含むテスト用Excelファイルを作成する"""
    file_path = tmp_path / "test_large.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "LargeSheet"

    # ヘッダー
    ws['A1'] = "ID"
    ws['B1'] = "質問"
    ws['C1'] = "回答"

    # 100行のデータを生成
    for i in range(2, 102):
        ws[f'A{i}'] = i - 1
        ws[f'B{i}'] = f"質問{i-1}: これはテスト用の質問です。"
        ws[f'C{i}'] = f"回答{i-1}: これはテスト用の回答です。"

    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def empty_excel_file(tmp_path):
    """空のテスト用Excelファイルを作成する"""
    file_path = tmp_path / "test_empty.xlsx"

    wb = Workbook()
    wb.save(file_path)
    return str(file_path)
