"""Excel解析サービス"""

import os
from typing import Dict, List, Optional, Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string


class ExcelParser:
    """Excelファイルを解析するクラス"""

    def parse_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Excelファイルを解析してシート情報を取得する

        Args:
            file_path: Excelファイルのパス

        Returns:
            ファイル情報とシート情報の辞書

        Raises:
            FileNotFoundError: ファイルが存在しない場合
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")

        wb = load_workbook(file_path, data_only=True)
        sheets = []

        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            sheet_info = {
                'name': sheet_name,
                'index': idx,
                'row_count': ws.max_row,
                'column_count': ws.max_column
            }
            sheets.append(sheet_info)

        return {
            'file_name': os.path.basename(file_path),
            'file_path': file_path,
            'sheets': sheets,
            'total_sheets': len(sheets)
        }

    def get_sheet_preview(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        start_row: int = 1,
        end_row: Optional[int] = None,
        start_column: int = 1,
        end_column: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        シートのプレビューデータを取得する

        Args:
            file_path: Excelファイルのパス
            sheet_name: シート名（未指定の場合は最初のシート）
            start_row: 開始行（1始まり）
            end_row: 終了行（未指定の場合はすべて）
            start_column: 開始列（1始まり）
            end_column: 終了列（未指定の場合はすべて）

        Returns:
            シートプレビューデータの辞書

        Raises:
            FileNotFoundError: ファイルが存在しない場合
            ValueError: シート名が存在しない場合
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")

        wb = load_workbook(file_path, data_only=False)

        # シートを取得
        if sheet_name is None:
            ws = wb.active
            sheet_name = ws.title
        else:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"シート '{sheet_name}' が見つかりません")
            ws = wb[sheet_name]

        # 範囲の決定
        if end_row is None:
            end_row = ws.max_row
        if end_column is None:
            end_column = ws.max_column

        # セルデータを取得
        cells = self._extract_cells(ws, start_row, end_row, start_column, end_column)

        return {
            'sheet_name': sheet_name,
            'cells': cells,
            'row_count': end_row - start_row + 1,
            'column_count': end_column - start_column + 1
        }

    def _extract_cells(
        self,
        ws: Worksheet,
        start_row: int,
        end_row: int,
        start_column: int,
        end_column: int
    ) -> List[Dict[str, Any]]:
        """
        ワークシートからセルデータを抽出する

        Args:
            ws: ワークシート
            start_row: 開始行
            end_row: 終了行
            start_column: 開始列
            end_column: 終了列

        Returns:
            セルデータのリスト
        """
        cells = []
        merged_cells_map = self._get_merged_cells_map(ws)

        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = ws.cell(row=row, column=col)
                cell_coord = f"{get_column_letter(col)}{row}"

                # セル結合情報を確認
                is_merged = cell_coord in merged_cells_map
                merge_range = merged_cells_map.get(cell_coord)

                cell_data = {
                    'row': row,
                    'column': col,
                    'value': cell.value,
                    'formatted_value': str(cell.value) if cell.value is not None else None,
                    'is_merged': is_merged,
                    'merge_range': merge_range
                }
                cells.append(cell_data)

        return cells

    def _get_merged_cells_map(self, ws: Worksheet) -> Dict[str, str]:
        """
        セル結合情報のマップを取得する

        Args:
            ws: ワークシート

        Returns:
            セル座標をキー、結合範囲を値とする辞書
        """
        merged_cells_map = {}

        for merged_range in ws.merged_cells.ranges:
            range_str = str(merged_range)
            # 結合範囲内のすべてのセルをマップに追加
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cell_coord = f"{get_column_letter(col)}{row}"
                    merged_cells_map[cell_coord] = range_str

        return merged_cells_map

    def extract_qa(
        self,
        file_path: str,
        sheet_name: str,
        start_row: int,
        end_row: int,
        question_column: int,
        answer_column: int,
        department_column: Optional[int] = None,
        skip_header_rows: int = 1
    ) -> Dict[str, Any]:
        """
        ExcelシートからQ/Aを抽出する

        Args:
            file_path: Excelファイルのパス
            sheet_name: シート名
            start_row: 開始行（1始まり）
            end_row: 終了行
            question_column: 質問列番号（1始まり）
            answer_column: 回答列番号（1始まり）
            department_column: 担当部門列番号（1始まり、任意）
            skip_header_rows: スキップするヘッダー行数

        Returns:
            抽出されたQ/Aデータの辞書

        Raises:
            FileNotFoundError: ファイルが存在しない場合
            ValueError: シート名が存在しない場合
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")

        wb = load_workbook(file_path, data_only=True)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"シート '{sheet_name}' が見つかりません")

        ws = wb[sheet_name]

        # Q/Aアイテムを抽出
        qa_items = []
        actual_start_row = start_row + skip_header_rows

        for row in range(actual_start_row, end_row + 1):
            # 質問セルの値を取得
            question_cell = ws.cell(row=row, column=question_column)
            question = str(question_cell.value).strip() if question_cell.value else ""

            # 質問が空の場合はスキップ
            if not question:
                continue

            # 回答セルの値を取得
            answer_cell = ws.cell(row=row, column=answer_column)
            answer = str(answer_cell.value).strip() if answer_cell.value else None

            # 部門セルの値を取得（指定されている場合）
            department = None
            if department_column is not None:
                dept_cell = ws.cell(row=row, column=department_column)
                department = str(dept_cell.value).strip() if dept_cell.value else None

            qa_item = {
                'row_number': row,
                'question': question,
                'answer': answer,
                'department': department
            }
            qa_items.append(qa_item)

        # 抽出元の範囲を計算
        question_col_letter = get_column_letter(question_column)
        answer_col_letter = get_column_letter(answer_column)

        # 最も左と最も右の列を決定
        columns = [question_column, answer_column]
        if department_column is not None:
            columns.append(department_column)
        min_col = min(columns)
        max_col = max(columns)

        min_col_letter = get_column_letter(min_col)
        max_col_letter = get_column_letter(max_col)
        source_range = f"{min_col_letter}{start_row}:{max_col_letter}{end_row}"

        return {
            'file_path': file_path,
            'sheet_name': sheet_name,
            'source_range': source_range,
            'items': qa_items,
            'total_items': len(qa_items)
        }
